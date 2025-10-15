#!/usr/bin/env python3
"""
Team Optimizer - Asymmetric Swap Algorithm (FIXED)
FIX: Η στήλη ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ χρησιμοποιεί 'Ν'/'Ο' (όχι 'Ν'/'O')
"""
import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
import io


@dataclass
class Student:
    """Δεδομένα μαθητή"""
    name: str
    choice: int
    gender: str
    greek_knowledge: str
    friends: List[str]
    locked: bool


class TeamOptimizer:
    """Asymmetric swap optimizer"""
    
    def __init__(self):
        self.students: Dict[str, Student] = {}
        self.teams: Dict[str, List[str]] = {}
        self.target_ep3 = 3
        self.target_gender = 4
        self.target_greek = 4
        
    def load_from_excel(self, file_bytes: bytes) -> None:
        """Διάβασμα completed Excel - FIX: Δεδομένα από ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ/SINGLE"""
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        
        print("\n🔍 DEBUG: Starting Excel load...")
        
        # ΒΗΜΑ 1: Διάβασε δεδομένα μαθητών από ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ
        if 'ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ' in wb.sheetnames:
            print("\n📄 Loading student data from ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ...")
            self._load_from_kategoriopoihsh(wb['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ'])
        
        # ΒΗΜΑ 2: Διάβασε δεδομένα από SINGLE
        if 'SINGLE' in wb.sheetnames:
            print("\n📄 Loading student data from SINGLE...")
            self._load_from_single(wb['SINGLE'])
        
        print(f"\n✅ Total students loaded: {len(self.students)}")
        
        # ΒΗΜΑ 3: Διάβασε team assignments από Α1, Α2, etc
        print("\n📄 Loading team assignments...")
        for sheet_name in wb.sheetnames:
            if sheet_name in ['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ', 'SINGLE', 'SWAP_SUGGESTIONS', 
                              'ΑΝΤΑΛΛΑΓΕΣ_ΑΝΑ_ΤΜΗΜΑ']:
                continue
            
            sheet = wb[sheet_name]
            headers = self._parse_headers(sheet)
            
            if 'ΟΝΟΜΑ' not in headers:
                continue
            
            self.teams[sheet_name] = []
            
            for row_idx in range(2, sheet.max_row + 1):
                name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
                if name and name in self.students:
                    self.teams[sheet_name].append(name)
            
            print(f"  ✅ {sheet_name}: {len(self.teams[sheet_name])} students")
        
        print(f"\n✅ Total teams: {len(self.teams)}\n")
        wb.close()
    
    def _load_from_kategoriopoihsh(self, sheet) -> None:
        """Διάβασμα δυάδων από ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ sheet"""
        headers = self._parse_headers(sheet)
        
        required = ['ΜΑΘΗΤΗΣΑ', 'ΜΑΘΗΤΗΣΒ', 'ΚΑΤΗΓΟΡΙΑΔΥΑΔΑΣ', 'ΕΠΙΔΟΣΗ']
        missing = [h for h in required if h not in headers]
        if missing:
            print(f"  ⚠️  Missing headers in ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ: {missing}")
            return
        
        pairs_loaded = 0
        
        for row_idx in range(2, sheet.max_row + 1):
            name_a = self._get_cell_value(sheet, row_idx, headers.get('ΜΑΘΗΤΗΣΑ'))
            name_b = self._get_cell_value(sheet, row_idx, headers.get('ΜΑΘΗΤΗΣΒ'))
            category = self._get_cell_value(sheet, row_idx, headers.get('ΚΑΤΗΓΟΡΙΑΔΥΑΔΑΣ'))
            epidosh_raw = self._get_cell_value(sheet, row_idx, headers.get('ΕΠΙΔΟΣΗ'))
            locked_val = self._get_cell_value(sheet, row_idx, headers.get('LOCKED'))
            
            if not name_a or not name_b or not category:
                continue
            
            # Parse επίδοση (format: "1,3" ή "2,2")
            epidosh_a, epidosh_b = 1, 1
            if ',' in epidosh_raw:
                parts = epidosh_raw.split(',')
                try:
                    epidosh_a = int(parts[0].strip())
                    epidosh_b = int(parts[1].strip())
                except:
                    pass
            
            # Parse φύλο και γλώσσα από category
            # Format: "όχι Καλή Γνώση (Αγόρια)" ή "Μικτή Γνώσης (Κορίτσια)"
            gender_a = gender_b = 'Α'
            greek_a = greek_b = 'Ν'
            
            if 'Αγόρια' in category or 'Αγόρ' in category:
                gender_a = gender_b = 'Α'
            elif 'Κορίτσια' in category or 'Κορίτ' in category:
                gender_a = gender_b = 'Κ'
            
            if 'όχι Καλή Γνώση' in category or 'όχι καλή' in category.lower():
                greek_a = greek_b = 'Ο'
            elif 'Καλή Γνώση' in category or 'Καλή γνώση' in category:
                greek_a = greek_b = 'Ν'
            elif 'Μικτή' in category or 'μικτή' in category.lower():
                # Μικτή - χρειάζεται επιπλέον λογική, default Ν
                greek_a = greek_b = 'Ν'
            
            is_locked = (locked_val == 'LOCKED')
            
            # Store students
            if name_a not in self.students:
                self.students[name_a] = Student(
                    name=name_a,
                    choice=epidosh_a,
                    gender=gender_a,
                    greek_knowledge=greek_a,
                    friends=[name_b],
                    locked=is_locked
                )
            
            if name_b not in self.students:
                self.students[name_b] = Student(
                    name=name_b,
                    choice=epidosh_b,
                    gender=gender_b,
                    greek_knowledge=greek_b,
                    friends=[name_a],
                    locked=is_locked
                )
            
            pairs_loaded += 1
        
        print(f"  ✅ Loaded {pairs_loaded} pairs ({pairs_loaded * 2} students)")
    
    def _load_from_single(self, sheet) -> None:
        """Διάβασμα μονών μαθητών από SINGLE sheet"""
        headers = self._parse_headers(sheet)
        
        required = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ']
        missing = [h for h in required if h not in headers]
        if missing:
            print(f"  ⚠️  Missing headers in SINGLE: {missing}")
            return
        
        singles_loaded = 0
        
        for row_idx in range(2, sheet.max_row + 1):
            name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
            if not name:
                continue
            
            # Αν ήδη φορτώθηκε από ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ, skip
            if name in self.students:
                continue
            
            gender_col = headers.get('ΦΥΛΟ') or headers.get('ΦΥΛΟ')
            greek_col = (headers.get('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ') or 
                        headers.get('ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ') or
                        headers.get('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ'))
            epidosh_col = headers.get('ΕΠΙΔΟΣΗ') or headers.get('ΕΠΙΔΟΣΗ')
            locked_col = headers.get('LOCKED')
            
            gender = self._get_cell_value(sheet, row_idx, gender_col, 'Α')
            
            # Greek knowledge
            raw_greek = sheet.cell(row_idx, greek_col).value if greek_col else 'Ν'
            if raw_greek and str(raw_greek).strip().upper().startswith('Ν'):
                greek = 'Ν'
            elif raw_greek and str(raw_greek).strip().upper().startswith('Ο'):
                greek = 'Ο'
            else:
                greek = 'Ν'
            
            # Επίδοση
            raw_epidosh = sheet.cell(row_idx, epidosh_col).value if epidosh_col else 1
            try:
                epidosh = int(raw_epidosh) if raw_epidosh else 1
            except:
                epidosh = 1
            
            locked_val = self._get_cell_value(sheet, row_idx, locked_col)
            is_locked = (locked_val == 'LOCKED' or locked_val == 'OΧΙ')
            
            self.students[name] = Student(
                name=name,
                choice=epidosh,
                gender=gender,
                greek_knowledge=greek,
                friends=[],
                locked=is_locked
            )
            
            singles_loaded += 1
        
        print(f"  ✅ Loaded {singles_loaded} single students")
    
    def _parse_headers(self, sheet) -> Dict[str, int]:
        """FIX: Normalization headers χωρίς να αφαιρούμε underscores"""
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                # Κρατάμε το original header
                raw_header = str(cell.value).strip()
                headers[raw_header] = col_idx
                
                # Και normalized version (για backward compatibility)
                normalized = raw_header.upper().replace(' ', '').replace('_', '')
                headers[normalized] = col_idx
        return headers
    
    def _get_cell_value(self, sheet, row: int, col: int, default=''):
        if col is None:
            return default
        val = sheet.cell(row, col).value
        return str(val).strip() if val is not None else default
    
    def _parse_friends(self, friends_str: str) -> List[str]:
        if not friends_str:
            return []
        return [f.strip() for f in friends_str.split(',') if f.strip()]
    
    def calculate_spreads(self) -> Dict[str, int]:
        """Υπολογισμός spreads"""
        stats = self._get_team_stats()
        if not stats:
            return {'ep3': 0, 'boys': 0, 'girls': 0, 'greek_yes': 0}
        
        ep3_vals = [s['ep3'] for s in stats.values()]
        boys_vals = [s['boys'] for s in stats.values()]
        girls_vals = [s['girls'] for s in stats.values()]
        greek_yes_vals = [s['greek_yes'] for s in stats.values()]
        
        return {
            'ep3': max(ep3_vals) - min(ep3_vals),
            'boys': max(boys_vals) - min(boys_vals),
            'girls': max(girls_vals) - min(girls_vals),
            'greek_yes': max(greek_yes_vals) - min(greek_yes_vals)
        }
    
    def _get_team_stats(self) -> Dict:
        """FIX: Διορθωμένη μέτρηση γλώσσας"""
        stats = {}
        for team_name, student_names in self.teams.items():
            boys = girls = greek_yes = greek_no = ep1 = ep2 = ep3 = 0
            
            for name in student_names:
                if name not in self.students:
                    continue
                s = self.students[name]
                
                # Gender
                if s.gender == 'Α':
                    boys += 1
                elif s.gender == 'Κ':
                    girls += 1
                
                # FIX: Greek knowledge - ελέγχουμε για 'Ν' (ΝΑΙ)
                if s.greek_knowledge == 'Ν':
                    greek_yes += 1
                elif s.greek_knowledge == 'Ο':
                    greek_no += 1
                
                # Choice
                if s.choice == 1:
                    ep1 += 1
                elif s.choice == 2:
                    ep2 += 1
                elif s.choice == 3:
                    ep3 += 1
            
            stats[team_name] = {
                'boys': boys, 'girls': girls,
                'greek_yes': greek_yes, 'greek_no': greek_no,
                'ep1': ep1, 'ep2': ep2, 'ep3': ep3
            }
        
        return stats
    
    def optimize(self, max_iterations: int = 100) -> Tuple[List[Dict], Dict]:
        """Asymmetric optimization"""
        applied_swaps = []
        
        for iteration in range(max_iterations):
            spreads = self.calculate_spreads()
            
            if (spreads['ep3'] <= self.target_ep3 and
                spreads['boys'] <= self.target_gender and
                spreads['girls'] <= self.target_gender and
                spreads['greek_yes'] <= self.target_greek):
                break
            
            stats = self._get_team_stats()
            ep3_counts = {team: stats[team]['ep3'] for team in stats.keys()}
            
            max_team = max(ep3_counts.items(), key=lambda x: x[1])[0]
            min_team = min(ep3_counts.items(), key=lambda x: x[1])[0]
            
            if ep3_counts[max_team] - ep3_counts[min_team] <= self.target_ep3:
                break
            
            all_swaps = self._generate_asymmetric_swaps(max_team, min_team)
            
            if not all_swaps:
                break
            
            best_swap = self._select_best_swap(all_swaps)
            
            if not best_swap:
                break
            
            self._apply_swap(best_swap)
            applied_swaps.append(best_swap)
        
        final_spreads = self.calculate_spreads()
        return applied_swaps, final_spreads
    
    def _generate_asymmetric_swaps(self, max_team: str, min_team: str) -> List[Dict]:
        """Γέννηση asymmetric swaps με 8 priorities"""
        swaps = []
        
        max_solos_ep3 = self._get_solos_with_ep3(max_team)
        max_pairs_ep3 = self._get_pairs_with_ep3(max_team)
        min_solos_non_ep3 = self._get_solos_without_ep3(min_team)
        min_pairs_non_ep3 = self._get_pairs_without_ep3(min_team)
        
        # P1: Solo(ep3) ↔ Solo(ep1/2), ίδιο φύλο+γλώσσα
        for solo_max in max_solos_ep3:
            for solo_min in min_solos_non_ep3:
                if (solo_max['student'].gender == solo_min['student'].gender and
                    solo_max['student'].greek_knowledge == solo_min['student'].greek_knowledge):
                    
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [solo_max['name']],
                        min_team, [solo_min['name']]
                    )
                    
                    if improvement['improves']:
                        swaps.append({
                            'type': 'Solo(ep3)↔Solo(ep1/2)-P1',
                            'from_team': max_team,
                            'students_out': [solo_max['name']],
                            'to_team': min_team,
                            'students_in': [solo_min['name']],
                            'improvement': improvement,
                            'priority': 1
                        })
        
        # P2: Δυάδα(ep3) ↔ Δυάδα(ep1/2), ίδιο φύλο+γλώσσα
        for pair_max in max_pairs_ep3:
            for pair_min in min_pairs_non_ep3:
                ep3_count_max = sum(1 for s in [pair_max['student_a'], pair_max['student_b']] if s.choice == 3)
                ep3_count_min = sum(1 for s in [pair_min['student_a'], pair_min['student_b']] if s.choice == 3)
                
                if ep3_count_max <= ep3_count_min:
                    continue
                
                genders_max = {pair_max['student_a'].gender, pair_max['student_b'].gender}
                genders_min = {pair_min['student_a'].gender, pair_min['student_b'].gender}
                greeks_max = {pair_max['student_a'].greek_knowledge, pair_max['student_b'].greek_knowledge}
                greeks_min = {pair_min['student_a'].greek_knowledge, pair_min['student_b'].greek_knowledge}
                
                if (len(genders_max) == 1 and len(genders_min) == 1 and genders_max == genders_min and
                    len(greeks_max) == 1 and len(greeks_min) == 1 and greeks_max == greeks_min):
                    
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [pair_max['name_a'], pair_max['name_b']],
                        min_team, [pair_min['name_a'], pair_min['name_b']]
                    )
                    
                    if improvement['improves']:
                        swaps.append({
                            'type': f"Δυάδα({pair_max['ep_combo']})↔Δυάδα({pair_min['ep_combo']})-P2",
                            'from_team': max_team,
                            'students_out': [pair_max['name_a'], pair_max['name_b']],
                            'to_team': min_team,
                            'students_in': [pair_min['name_a'], pair_min['name_b']],
                            'improvement': improvement,
                            'priority': 2
                        })
        
        # P3-8: Χαλάρωση περιορισμών (όπως πριν)
        for solo_max in max_solos_ep3:
            for solo_min in min_solos_non_ep3:
                if solo_max['student'].gender == solo_min['student'].gender:
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [solo_max['name']],
                        min_team, [solo_min['name']]
                    )
                    if improvement['improves']:
                        swaps.append({
                            'type': 'Solo(ep3)↔Solo(ep1/2)-P3',
                            'from_team': max_team,
                            'students_out': [solo_max['name']],
                            'to_team': min_team,
                            'students_in': [solo_min['name']],
                            'improvement': improvement,
                            'priority': 3
                        })
        
        for pair_max in max_pairs_ep3:
            for pair_min in min_pairs_non_ep3:
                ep3_count_max = sum(1 for s in [pair_max['student_a'], pair_max['student_b']] if s.choice == 3)
                ep3_count_min = sum(1 for s in [pair_min['student_a'], pair_min['student_b']] if s.choice == 3)
                if ep3_count_max <= ep3_count_min:
                    continue
                genders_max = {pair_max['student_a'].gender, pair_max['student_b'].gender}
                genders_min = {pair_min['student_a'].gender, pair_min['student_b'].gender}
                if len(genders_max) == 1 and len(genders_min) == 1 and genders_max == genders_min:
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [pair_max['name_a'], pair_max['name_b']],
                        min_team, [pair_min['name_a'], pair_min['name_b']]
                    )
                    if improvement['improves']:
                        swaps.append({
                            'type': f"Δυάδα({pair_max['ep_combo']})↔Δυάδα({pair_min['ep_combo']})-P4",
                            'from_team': max_team,
                            'students_out': [pair_max['name_a'], pair_max['name_b']],
                            'to_team': min_team,
                            'students_in': [pair_min['name_a'], pair_min['name_b']],
                            'improvement': improvement,
                            'priority': 4
                        })
        
        for solo_max in max_solos_ep3:
            for solo_min in min_solos_non_ep3:
                if solo_max['student'].greek_knowledge == solo_min['student'].greek_knowledge:
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [solo_max['name']],
                        min_team, [solo_min['name']]
                    )
                    if improvement['improves']:
                        swaps.append({
                            'type': 'Solo(ep3)↔Solo(ep1/2)-P5',
                            'from_team': max_team,
                            'students_out': [solo_max['name']],
                            'to_team': min_team,
                            'students_in': [solo_min['name']],
                            'improvement': improvement,
                            'priority': 5
                        })
        
        for pair_max in max_pairs_ep3:
            for pair_min in min_pairs_non_ep3:
                ep3_count_max = sum(1 for s in [pair_max['student_a'], pair_max['student_b']] if s.choice == 3)
                ep3_count_min = sum(1 for s in [pair_min['student_a'], pair_min['student_b']] if s.choice == 3)
                if ep3_count_max <= ep3_count_min:
                    continue
                greeks_max = {pair_max['student_a'].greek_knowledge, pair_max['student_b'].greek_knowledge}
                greeks_min = {pair_min['student_a'].greek_knowledge, pair_min['student_b'].greek_knowledge}
                if len(greeks_max) == 1 and len(greeks_min) == 1 and greeks_max == greeks_min:
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [pair_max['name_a'], pair_max['name_b']],
                        min_team, [pair_min['name_a'], pair_min['name_b']]
                    )
                    if improvement['improves']:
                        swaps.append({
                            'type': f"Δυάδα({pair_max['ep_combo']})↔Δυάδα({pair_min['ep_combo']})-P6",
                            'from_team': max_team,
                            'students_out': [pair_max['name_a'], pair_max['name_b']],
                            'to_team': min_team,
                            'students_in': [pair_min['name_a'], pair_min['name_b']],
                            'improvement': improvement,
                            'priority': 6
                        })
        
        for solo_max in max_solos_ep3:
            for solo_min in min_solos_non_ep3:
                improvement = self._calc_asymmetric_improvement(
                    max_team, [solo_max['name']],
                    min_team, [solo_min['name']]
                )
                if improvement['improves']:
                    swaps.append({
                        'type': f"Solo({solo_max['student'].choice})↔Solo({solo_min['student'].choice})-P7",
                        'from_team': max_team,
                        'students_out': [solo_max['name']],
                        'to_team': min_team,
                        'students_in': [solo_min['name']],
                        'improvement': improvement,
                        'priority': 7
                    })
        
        for pair_max in max_pairs_ep3:
            for pair_min in min_pairs_non_ep3:
                ep3_count_max = sum(1 for s in [pair_max['student_a'], pair_max['student_b']] if s.choice == 3)
                ep3_count_min = sum(1 for s in [pair_min['student_a'], pair_min['student_b']] if s.choice == 3)
                if ep3_count_max <= ep3_count_min:
                    continue
                improvement = self._calc_asymmetric_improvement(
                    max_team, [pair_max['name_a'], pair_max['name_b']],
                    min_team, [pair_min['name_a'], pair_min['name_b']]
                )
                if improvement['improves']:
                    swaps.append({
                        'type': f"Δυάδα({pair_max['ep_combo']})↔Δυάδα({pair_min['ep_combo']})-P8",
                        'from_team': max_team,
                        'students_out': [pair_max['name_a'], pair_max['name_b']],
                        'to_team': min_team,
                        'students_in': [pair_min['name_a'], pair_min['name_b']],
                        'improvement': improvement,
                        'priority': 8
                    })
        
        return swaps
    
    def _get_solos_with_ep3(self, team_name: str) -> List[Dict]:
        solos = []
        student_names = self.teams[team_name]
        for name in student_names:
            if name not in self.students:
                continue
            student = self.students[name]
            if student.locked or student.choice != 3:
                continue
            has_friend = any(f in student_names for f in student.friends)
            if not has_friend:
                solos.append({'name': name, 'student': student})
        return solos
    
    def _get_pairs_with_ep3(self, team_name: str) -> List[Dict]:
        pairs = []
        processed = set()
        student_names = self.teams[team_name]
        for name_a in student_names:
            if name_a in processed or name_a not in self.students:
                continue
            student_a = self.students[name_a]
            if student_a.locked:
                continue
            for name_b in student_names:
                if name_b == name_a or name_b in processed or name_b not in self.students:
                    continue
                student_b = self.students[name_b]
                if student_b.locked:
                    continue
                if name_b in student_a.friends or name_a in student_b.friends:
                    if student_a.choice == 3 or student_b.choice == 3:
                        pairs.append({
                            'name_a': name_a, 'name_b': name_b,
                            'student_a': student_a, 'student_b': student_b,
                            'ep_combo': f"{student_a.choice},{student_b.choice}"
                        })
                        processed.add(name_a)
                        processed.add(name_b)
                        break
        return pairs
    
    def _get_solos_without_ep3(self, team_name: str) -> List[Dict]:
        solos = []
        student_names = self.teams[team_name]
        for name in student_names:
            if name not in self.students:
                continue
            student = self.students[name]
            if student.locked or student.choice == 3:
                continue
            has_friend = any(f in student_names for f in student.friends)
            if not has_friend:
                solos.append({'name': name, 'student': student})
        return solos
    
    def _get_pairs_without_ep3(self, team_name: str) -> List[Dict]:
        pairs = []
        processed = set()
        student_names = self.teams[team_name]
        for name_a in student_names:
            if name_a in processed or name_a not in self.students:
                continue
            student_a = self.students[name_a]
            if student_a.locked:
                continue
            for name_b in student_names:
                if name_b == name_a or name_b in processed or name_b not in self.students:
                    continue
                student_b = self.students[name_b]
                if student_b.locked:
                    continue
                if name_b in student_a.friends or name_a in student_b.friends:
                    pairs.append({
                        'name_a': name_a, 'name_b': name_b,
                        'student_a': student_a, 'student_b': student_b,
                        'ep_combo': f"{student_a.choice},{student_b.choice}"
                    })
                    processed.add(name_a)
                    processed.add(name_b)
                    break
        return pairs
    
    def _calc_asymmetric_improvement(self, team_high: str, names_out: List[str],
                                      team_low: str, names_in: List[str]) -> Dict:
        """FIX: Διορθωμένος υπολογισμός με 'Ν'/'Ο'"""
        stats_before = self._get_team_stats()
        stats_after = {k: v.copy() for k, v in stats_before.items()}
        
        for name in names_out:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_high]['ep3'] -= 1
                if s.gender == 'Α': stats_after[team_high]['boys'] -= 1
                elif s.gender == 'Κ': stats_after[team_high]['girls'] -= 1
                if s.greek_knowledge == 'Ν': stats_after[team_high]['greek_yes'] -= 1
        
        for name in names_in:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_high]['ep3'] += 1
                if s.gender == 'Α': stats_after[team_high]['boys'] += 1
                elif s.gender == 'Κ': stats_after[team_high]['girls'] += 1
                if s.greek_knowledge == 'Ν': stats_after[team_high]['greek_yes'] += 1
        
        for name in names_in:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_low]['ep3'] -= 1
                if s.gender == 'Α': stats_after[team_low]['boys'] -= 1
                elif s.gender == 'Κ': stats_after[team_low]['girls'] -= 1
                if s.greek_knowledge == 'Ν': stats_after[team_low]['greek_yes'] -= 1
        
        for name in names_out:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_low]['ep3'] += 1
                if s.gender == 'Α': stats_after[team_low]['boys'] += 1
                elif s.gender == 'Κ': stats_after[team_low]['girls'] += 1
                if s.greek_knowledge == 'Ν': stats_after[team_low]['greek_yes'] += 1
        
        ep3_before = max(s['ep3'] for s in stats_before.values()) - min(s['ep3'] for s in stats_before.values())
        ep3_after = max(s['ep3'] for s in stats_after.values()) - min(s['ep3'] for s in stats_after.values())
        
        boys_before = max(s['boys'] for s in stats_before.values()) - min(s['boys'] for s in stats_before.values())
        boys_after = max(s['boys'] for s in stats_after.values()) - min(s['boys'] for s in stats_after.values())
        
        girls_before = max(s['girls'] for s in stats_before.values()) - min(s['girls'] for s in stats_before.values())
        girls_after = max(s['girls'] for s in stats_after.values()) - min(s['girls'] for s in stats_after.values())
        
        greek_before = max(s['greek_yes'] for s in stats_before.values()) - min(s['greek_yes'] for s in stats_before.values())
        greek_after = max(s['greek_yes'] for s in stats_after.values()) - min(s['greek_yes'] for s in stats_after.values())
        
        delta_ep3 = ep3_before - ep3_after
        delta_boys = boys_before - boys_after
        delta_girls = girls_before - girls_after
        delta_greek = greek_before - greek_after
        
        improves = delta_ep3 > 0 or (delta_ep3 == 0 and (delta_boys > 0 or delta_girls > 0 or delta_greek > 0))
        
        return {
            'improves': improves,
            'delta_ep3': delta_ep3,
            'delta_boys': delta_boys,
            'delta_girls': delta_girls,
            'delta_greek': delta_greek,
            'ep3_before': ep3_before,
            'ep3_after': ep3_after
        }
    
    def _select_best_swap(self, swaps: List[Dict]) -> Optional[Dict]:
        if not swaps:
            return None
        
        swaps.sort(
            key=lambda x: (
                -x['improvement']['delta_ep3'],
                -(x['improvement']['delta_boys'] + x['improvement']['delta_girls']),
                -x['improvement']['delta_greek'],
                x['priority']
            )
        )
        
        return swaps[0]
    
    def _apply_swap(self, swap: Dict) -> None:
        from_team = swap['from_team']
        to_team = swap['to_team']
        students_out = swap['students_out']
        students_in = swap['students_in']
        
        for name in students_out:
            if name in self.teams[from_team]:
                self.teams[from_team].remove(name)
        
        for name in students_in:
            if name in self.teams[to_team]:
                self.teams[to_team].remove(name)
        
        for name in students_out:
            self.teams[to_team].append(name)
        
        for name in students_in:
            self.teams[from_team].append(name)
    
    def export_to_excel(self, applied_swaps: List[Dict], final_spreads: Dict) -> bytes:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        for team_name in sorted(self.teams.keys()):
            self._create_team_sheet(wb, team_name)
        
        self._create_statistics_sheet(wb, final_spreads)
        self._create_swaps_log_sheet(wb, applied_swaps)
        
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)
        
        return output.getvalue()
    
    def _create_team_sheet(self, wb, team_name: str) -> None:
        sheet = wb.create_sheet(team_name)
        
        headers = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΦΙΛΟΙ']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDEBF7', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx = 2
        for name in sorted(self.teams[team_name]):
            if name not in self.students:
                continue
            
            student = self.students[name]
            sheet.cell(row_idx, 1).value = student.name
            sheet.cell(row_idx, 2).value = student.gender
            sheet.cell(row_idx, 3).value = student.greek_knowledge
            sheet.cell(row_idx, 4).value = student.choice
            sheet.cell(row_idx, 5).value = ', '.join(student.friends)
            
            for col in range(1, 6):
                sheet.cell(row_idx, col).alignment = Alignment(
                    horizontal='left' if col in [1,5] else 'center', 
                    vertical='center'
                )
            
            row_idx += 1
        
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 40
    
    def _create_statistics_sheet(self, wb, spreads: Dict) -> None:
        sheet = wb.create_sheet('ΒΕΛΤΙΩΜΕΝΗ_ΣΤΑΤΙΣΤΙΚΗ')
        
        headers = ['Τμήμα', 'Σύνολο', 'Αγόρια', 'Κορίτσια', 
                   'Γνώση (ΝΑΙ)', 'Γνώση (ΟΧΙ)', 'Επ1', 'Επ2', 'Επ3']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='C6E0B4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        stats = self._get_team_stats()
        row_idx = 2
        for team_name in sorted(self.teams.keys()):
            if team_name not in stats:
                continue
            s = stats[team_name]
            
            sheet.cell(row_idx, 1).value = team_name
            sheet.cell(row_idx, 2).value = len(self.teams[team_name])
            sheet.cell(row_idx, 3).value = s['boys']
            sheet.cell(row_idx, 4).value = s['girls']
            sheet.cell(row_idx, 5).value = s['greek_yes']
            sheet.cell(row_idx, 6).value = s['greek_no']
            sheet.cell(row_idx, 7).value = s['ep1']
            sheet.cell(row_idx, 8).value = s['ep2']
            sheet.cell(row_idx, 9).value = s['ep3']
            
            for col in range(1, 10):
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            row_idx += 1
        
        row_idx += 2
        sheet.cell(row_idx, 1).value = 'ΤΕΛΙΚΑ SPREADS'
        sheet.cell(row_idx, 1).font = Font(bold=True, size=12)
        row_idx += 1
        
        summary_headers = ['Μετρική', 'Spread', 'Στόχος', 'Status']
        for col_idx, header in enumerate(summary_headers, start=1):
            cell = sheet.cell(row_idx, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFF2CC', fill_type='solid')
        row_idx += 1
        
        summary_data = [
            ('Spread Επίδοσης 3', spreads['ep3'], '≤ 3', '✅' if spreads['ep3'] <= 3 else '❌'),
            ('Spread Αγοριών', spreads['boys'], '≤ 4', '✅' if spreads['boys'] <= 4 else '❌'),
            ('Spread Κοριτσιών', spreads['girls'], '≤ 4', '✅' if spreads['girls'] <= 4 else '❌'),
            ('Spread Γνώσης', spreads['greek_yes'], '≤ 4', '✅' if spreads['greek_yes'] <= 4 else '❌')
        ]
        
        for label, value, target, status in summary_data:
            sheet.cell(row_idx, 1).value = label
            sheet.cell(row_idx, 2).value = value
            sheet.cell(row_idx, 3).value = target
            sheet.cell(row_idx, 4).value = status
            
            if '✅' in status:
                sheet.cell(row_idx, 2).fill = PatternFill(start_color='C6EFCE', fill_type='solid')
            else:
                sheet.cell(row_idx, 2).fill = PatternFill(start_color='FFC7CE', fill_type='solid')
            
            row_idx += 1
        
        for col in ['A', 'B', 'C', 'D']:
            sheet.column_dimensions[col].width = 20
    
    def _create_swaps_log_sheet(self, wb, swaps: List[Dict]) -> None:
        sheet = wb.create_sheet('ΕΦΑΡΜΟΣΜΕΝΑ_SWAPS')
        
        headers = ['#', 'Τύπος', 'Από Τμήμα', 'Μαθητές OUT (ep3)', 
                   'Προς Τμήμα', 'Μαθητές IN (ep1/2)', 'Δ_ep3', 'Δ_φύλου', 'Δ_γνώσης', 'Priority']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for idx, swap in enumerate(swaps, start=1):
            imp = swap['improvement']
            
            sheet.cell(idx + 1, 1).value = idx
            sheet.cell(idx + 1, 2).value = swap['type']
            sheet.cell(idx + 1, 3).value = swap['from_team']
            sheet.cell(idx + 1, 4).value = ', '.join(swap['students_out'])
            sheet.cell(idx + 1, 5).value = swap['to_team']
            sheet.cell(idx + 1, 6).value = ', '.join(swap['students_in'])
            sheet.cell(idx + 1, 7).value = f"+{imp['delta_ep3']}" if imp['delta_ep3'] > 0 else str(imp['delta_ep3'])
            sheet.cell(idx + 1, 8).value = f"+{imp['delta_boys'] + imp['delta_girls']}" if imp['delta_boys'] + imp['delta_girls'] > 0 else str(imp['delta_boys'] + imp['delta_girls'])
            sheet.cell(idx + 1, 9).value = f"+{imp['delta_greek']}" if imp['delta_greek'] > 0 else str(imp['delta_greek'])
            sheet.cell(idx + 1, 10).value = swap['priority']
            
            for col in range(1, 11):
                sheet.cell(idx + 1, col).alignment = Alignment(horizontal='center', vertical='center')
        
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 35
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 35
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 10


def main():
    st.set_page_config(
        page_title="Team Optimizer (FIXED)",
        page_icon="🎯",
        layout="wide"
    )
    
    st.title("🎯 Team Optimizer - FIXED Greek Knowledge Bug")
    st.markdown("---")
    
    with st.expander("🔧 FIX Details", expanded=False):
        st.markdown("""
        **Διόρθωση:**
        - ✅ Σωστή ανάγνωση στήλης `ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ`
        - ✅ Normalization: 'Ν' = ΝΑΙ, 'Ο' = ΟΧΙ
        - ✅ Σωστή μέτρηση στα statistics
        - ✅ Σωστή εξαγωγή στο Excel
        """)
    
    with st.expander("📖 Οδηγίες Χρήσης", expanded=False):
        st.markdown("""
        **Λογική Asymmetric Swaps:**
        - Τμήμα με **πολλά ep3** δίνει μαθητές με **επίδοση 3**
        - Τμήμα με **λίγα ep3** δίνει μαθητές με **επίδοση 1 ή 2**
        
        **Προτεραιότητες:**
        1. Solo(ep3) ↔ Solo(ep1/2) - Ίδιο φύλο + γλώσσα
        2. Δυάδα(ep3) ↔ Δυάδα(ep1/2) - Ίδιο φύλο + γλώσσα
        3-8. Χαλάρωση περιορισμών
        
        **Στόχοι:**
        - Spread Επίδοσης 3: ≤ 3 ✅
        - Spread Φύλου: ≤ 4 ✅
        - Spread Γνώσης: ≤ 4 ✅
        """)
    
    st.subheader("📥 Upload Completed Excel")
    completed_file = st.file_uploader(
        "Ανέβασε το STEP7_COMPLETED.xlsx",
        type=['xlsx'],
        key='completed'
    )
    
    if completed_file:
        st.success(f"✅ {completed_file.name}")
        
        if st.button("⚡ Εκτέλεση Optimization", type="primary", use_container_width=True):
            with st.spinner("🔄 Asymmetric swaps σε εξέλιξη..."):
                try:
                    optimizer = TeamOptimizer()
                    optimizer.load_from_excel(completed_file.read())
                    
                    # Debug: Εμφάνιση sample students
                    with st.expander("🔍 Debug: Sample Students", expanded=False):
                        sample_students = list(optimizer.students.items())[:5]
                        for name, student in sample_students:
                            st.text(f"{name}: Greek={student.greek_knowledge}, Gender={student.gender}, Choice={student.choice}")
                    
                    spreads_before = optimizer.calculate_spreads()
                    stats_before = optimizer._get_team_stats()
                    
                    st.info("📊 **ΠΡΙΝ την Βελτιστοποίηση:**")
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Spread Επ3", spreads_before['ep3'])
                    with col2:
                        st.metric("Spread Αγόρια", spreads_before['boys'])
                    with col3:
                        st.metric("Spread Κορίτσια", spreads_before['girls'])
                    with col4:
                        st.metric("Spread Γνώση", spreads_before['greek_yes'])
                    
                    # Debug stats
                    with st.expander("📊 Detailed Stats BEFORE", expanded=False):
                        for team, s in stats_before.items():
                            st.text(f"{team}: ΝΑΙ={s['greek_yes']}, ΟΧΙ={s['greek_no']}, EP3={s['ep3']}")
                    
                    # Optimization
                    applied_swaps, spreads_after = optimizer.optimize(max_iterations=100)
                    stats_after = optimizer._get_team_stats()
                    
                    st.markdown("---")
                    st.success("✅ **ΜΕΤΑ την Βελτιστοποίηση:**")
                    
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric(
                            "Spread Επ3", 
                            spreads_after['ep3'],
                            delta=-(spreads_before['ep3'] - spreads_after['ep3']),
                            delta_color="inverse"
                        )
                        if spreads_after['ep3'] <= 3:
                            st.success("✅ Στόχος επιτεύχθηκε!")
                        else:
                            st.warning(f"⚠️ Στόχος: ≤ 3")
                    
                    with col2:
                        st.metric(
                            "Spread Αγόρια",
                            spreads_after['boys'],
                            delta=-(spreads_before['boys'] - spreads_after['boys']),
                            delta_color="inverse"
                        )
                        if spreads_after['boys'] <= 4:
                            st.success("✅")
                        else:
                            st.warning("⚠️ ≤ 4")
                    
                    with col3:
                        st.metric(
                            "Spread Κορίτσια",
                            spreads_after['girls'],
                            delta=-(spreads_before['girls'] - spreads_after['girls']),
                            delta_color="inverse"
                        )
                        if spreads_after['girls'] <= 4:
                            st.success("✅")
                        else:
                            st.warning("⚠️ ≤ 4")
                    
                    with col4:
                        st.metric(
                            "Spread Γνώση",
                            spreads_after['greek_yes'],
                            delta=-(spreads_before['greek_yes'] - spreads_after['greek_yes']),
                            delta_color="inverse"
                        )
                        if spreads_after['greek_yes'] <= 4:
                            st.success("✅")
                        else:
                            st.warning("⚠️ ≤ 4")
                    
                    # Debug stats AFTER
                    with st.expander("📊 Detailed Stats AFTER", expanded=False):
                        for team, s in stats_after.items():
                            st.text(f"{team}: ΝΑΙ={s['greek_yes']}, ΟΧΙ={s['greek_no']}, EP3={s['ep3']}")
                    
                    st.markdown("---")
                    st.info(f"🔄 **Εφαρμόστηκαν {len(applied_swaps)} swaps συνολικά**")
                    
                    # Export
                    output_bytes = optimizer.export_to_excel(applied_swaps, spreads_after)
                    
                    st.download_button(
                        label="📥 Κατέβασε Βελτιωμένη Κατανομή",
                        data=output_bytes,
                        file_name="ΒΕΛΤΙΩΜΕΝΗ_ΚΑΤΑΝΟΜΗ_FIXED.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
                    
                    st.balloons()
                    
                except Exception as e:
                    st.error(f"❌ Σφάλμα: {str(e)}")
                    with st.expander("Λεπτομέρειες"):
                        import traceback
                        st.code(traceback.format_exc())
    else:
        st.info("👆 Ανέβασε το completed Excel για να ξεκινήσεις")
    
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: gray;'>"
        "Team Optimizer v2.1 FIXED | Greek Knowledge Bug Resolved ✅"
        "</div>",
        unsafe_allow_html=True
    )


if __name__ == '__main__':
    main()
